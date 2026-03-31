"""
application/vnd.openxmlformats-officedocument.spreadsheetml.sheet
csv_to_excel_formatter.py
─────────────────────────────────────────────────────────────────────────────
Generic, production-grade CSV → Professional Excel converter.
Works with any CSV file — no schema assumptions.

Usage:
    python csv_to_excel_formatter.py --input data.csv --output report.xlsx
    python csv_to_excel_formatter.py --input data.csv --output report.xlsx --title "Sales Report" --sheet "Q1 Data"
    python csv_to_excel_formatter.py --input data.csv --output report.xlsx --summary   # adds auto-summary sheet

Dependencies:
    pip install pandas openpyxl xlsxwriter
─────────────────────────────────────────────────────────────────────────────
"""

import argparse
import sys
from pathlib import Path
from datetime import datetime

import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.page import PageMargins
from openpyxl.drawing.image import Image


# ─── THEME ────────────────────────────────────────────────────────────────────

class Theme:
    """Centralised colour and font palette. Edit here to retheme the whole sheet."""

    # Core palette — deep navy / slate family
    HEADER_BG       = "1B2A4A"   # dark navy
    HEADER_FG       = "FFFFFF"   # white text
    ACCENT          = "2E75B6"   # medium blue (used for title bar, totals)
    ACCENT_LIGHT    = "D6E4F0"   # pale blue (alternate row tint)
    ROW_ALT         = "F0F4F8"   # subtle stripe
    ROW_DEFAULT     = "FFFFFF"
    TOTAL_BG        = "E8EEF4"   # totals row background
    TOTAL_FG        = "1B2A4A"   # totals row text
    TITLE_BG        = "1B2A4A"
    TITLE_FG        = "FFFFFF"
    BORDER_COLOR    = "B8C8D8"
    FONT_FAMILY     = "Aptos"       # Falls back gracefully if unavailable
    FONT_FAMILY_ALT = "Calibri"   # Fallback


# ─── STYLE HELPERS ────────────────────────────────────────────────────────────

def _border(style="thin", color=Theme.BORDER_COLOR):
    side = Side(style=style, color=color)
    return Border(left=side, right=side, top=side, bottom=side)

def _header_bottom_border():
    thick = Side(style="medium", color=Theme.ACCENT)
    thin  = Side(style="thin",   color=Theme.BORDER_COLOR)
    return Border(left=thin, right=thin, top=thin, bottom=thick)

def _fill(hex_color):
    return PatternFill(fill_type="solid", start_color=hex_color, end_color=hex_color)

def _font(bold=False, color="000000", size=11, italic=False, name=Theme.FONT_FAMILY):
    return Font(name=name, bold=bold, color=color, size=size, italic=italic)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ─── COLUMN INTELLIGENCE ──────────────────────────────────────────────────────

def _classify_column(series: pd.Series) -> str:
    """Detect the semantic type of a pandas Series to drive number formatting."""
    name_lower = series.name.lower() if isinstance(series.name, str) else ""
    dtype = series.dtype

    if pd.api.types.is_datetime64_any_dtype(dtype):
        return "date"
    if pd.api.types.is_bool_dtype(dtype):
        return "boolean"
    if pd.api.types.is_integer_dtype(dtype):
        # heuristic: looks like a year column?
        if "year" in name_lower and series.between(1900, 2100).all():
            return "year"
        return "integer"
    if pd.api.types.is_float_dtype(dtype):
        # heuristic: percentage column?
        if any(k in name_lower for k in ("pct", "percent", "rate", "ratio", "margin")):
            return "percent"
        # heuristic: currency / money column?
        if any(k in name_lower for k in ("price", "cost", "revenue", "sales", "amount",
                                          "salary", "income", "profit", "loss", "fee",
                                          "budget", "spend", "value", "total", "sum")):
            return "currency"
        return "float"
    return "text"


_FORMAT_MAP = {
    "currency": '#,##0.00',
    "integer":  '#,##0',
    "float":    '#,##0.00',
    "percent":  '0.00%',
    "date":     'DD-MMM-YYYY',
    "year":     '0',
    "boolean":  '@',
    "text":     '@',
}


# ─── OPTIMAL COLUMN WIDTH ─────────────────────────────────────────────────────

def _auto_width(ws, df: pd.DataFrame, offset_col: int = 1):
    """Calculate and apply optimal column widths based on content."""
    for idx, col in enumerate(df.columns, start=offset_col):
        col_letter = get_column_letter(idx)
        header_len = len(str(col))
        try:
            max_data_len = df[col].astype(str).str.len().max()
        except Exception:
            max_data_len = 10
        optimal = min(max(header_len, int(max_data_len or 10)) + 4, 50)
        ws.column_dimensions[col_letter].width = optimal


# ─── SUMMARY SHEET ────────────────────────────────────────────────────────────

def _build_summary_sheet(wb: Workbook, df: pd.DataFrame, source_name: str):
    """Append an auto-generated Data Summary sheet."""
    ws = wb.create_sheet("📊 Summary")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 35

    # Title
    ws.merge_cells("A1:B1")
    ws["A1"] = "Dataset Summary"
    ws["A1"].font = _font(bold=True, color=Theme.TITLE_FG, size=14)
    ws["A1"].fill = _fill(Theme.TITLE_BG)
    ws["A1"].alignment = _align("center")
    ws.row_dimensions[1].height = 32

    meta_rows = [
        ("Source File",     source_name),
        ("Generated On",    datetime.now().strftime("%d %b %Y, %H:%M")),
        ("Total Rows",      f"{len(df):,}"),
        ("Total Columns",   str(len(df.columns))),
        ("Numeric Columns", str(df.select_dtypes(include="number").shape[1])),
        ("Text Columns",    str(df.select_dtypes(include=["object", "string"]).shape[1])),
        ("Missing Values",  f"{df.isnull().sum().sum():,}"),
        ("Duplicate Rows",  f"{df.duplicated().sum():,}"),
    ]

    for r_idx, (label, value) in enumerate(meta_rows, start=2):
        label_cell = ws.cell(row=r_idx, column=1, value=label)
        value_cell = ws.cell(row=r_idx, column=2, value=value)
        bg = Theme.ROW_ALT if r_idx % 2 == 0 else Theme.ROW_DEFAULT
        for cell in (label_cell, value_cell):
            cell.fill = _fill(bg)
            cell.border = _border()
            cell.alignment = _align()
        label_cell.font = _font(bold=True, size=10)
        value_cell.font = _font(size=10)
        ws.row_dimensions[r_idx].height = 20

    # Per-column stats for numeric columns
    num_df = df.select_dtypes(include="number")
    if not num_df.empty:
        start_row = len(meta_rows) + 4
        ws.merge_cells(f"A{start_row}:B{start_row}")
        ws[f"A{start_row}"] = "Numeric Column Statistics"
        ws[f"A{start_row}"].font = _font(bold=True, color=Theme.TITLE_FG, size=11)
        ws[f"A{start_row}"].fill = _fill(Theme.ACCENT)
        ws[f"A{start_row}"].alignment = _align("center")
        ws.row_dimensions[start_row].height = 26

        stat_headers = ["Column", "Min", "Max", "Mean", "Nulls"]
        for c_idx, h in enumerate(stat_headers, start=1):
            cell = ws.cell(row=start_row + 1, column=c_idx, value=h)
            cell.font = _font(bold=True, color=Theme.HEADER_FG, size=10)
            cell.fill = _fill(Theme.HEADER_BG)
            cell.alignment = _align("center")
            cell.border = _border()

        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 18

        for r_idx, col in enumerate(num_df.columns, start=start_row + 2):
            vals = [
                col,
                round(float(num_df[col].min()), 2) if pd.notna(num_df[col].min()) else "N/A",
                round(float(num_df[col].max()), 2) if pd.notna(num_df[col].max()) else "N/A",
                round(float(num_df[col].mean()), 2) if pd.notna(num_df[col].mean()) else "N/A",
                int(num_df[col].isnull().sum()),
            ]
            bg = Theme.ROW_ALT if r_idx % 2 == 0 else Theme.ROW_DEFAULT
            for c_idx, v in enumerate(vals, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=v)
                cell.fill = _fill(bg)
                cell.border = _border()
                cell.alignment = _align("center" if c_idx > 1 else "left")
                cell.font = _font(size=10)
            ws.row_dimensions[r_idx].height = 18


# ─── MAIN DATA SHEET ──────────────────────────────────────────────────────────

def _build_data_sheet(
    wb: Workbook,
    df: pd.DataFrame,
    sheet_name: str,
    report_title: str,
    source_name: str,
):
    ws = wb.active
    ws.title = sheet_name
    ws.sheet_view.showGridLines = False

    # ── Title banner ─────────────────────────────────────────────────────────
    TITLE_ROW   = 1
    SUBTITLE_ROW = 2
    HEADER_ROW  = 4          # leave row 3 as breathing space
    DATA_START  = HEADER_ROW + 1
    n_cols      = len(df.columns)

    last_col_letter = get_column_letter(n_cols)

    ws.merge_cells(f"A{TITLE_ROW}:{last_col_letter}{TITLE_ROW}")
    title_cell = ws[f"A{TITLE_ROW}"]
    title_cell.value = report_title
    title_cell.font = _font(bold=True, color=Theme.TITLE_FG, size=16)
    title_cell.fill = _fill(Theme.TITLE_BG)
    title_cell.alignment = _align("center")
    ws.row_dimensions[TITLE_ROW].height = 38

    ws.merge_cells(f"A{SUBTITLE_ROW}:{last_col_letter}{SUBTITLE_ROW}")
    sub_cell = ws[f"A{SUBTITLE_ROW}"]
    sub_cell.value = (
        f"Source: {source_name}   |   "
        f"Rows: {len(df):,}   |   "
        f"Generated: {datetime.now().strftime('%d %b %Y, %H:%M')}"
    )
    sub_cell.font = _font(italic=True, color="AAAAAA", size=9)
    sub_cell.fill = _fill(Theme.TITLE_BG)
    sub_cell.alignment = _align("center")
    ws.row_dimensions[SUBTITLE_ROW].height = 18

    # ── Column type classification ────────────────────────────────────────────
    col_types = {col: _classify_column(df[col]) for col in df.columns}

    # ── Headers ───────────────────────────────────────────────────────────────
    ws.row_dimensions[3].height = 8        # visual spacer
    ws.row_dimensions[HEADER_ROW].height = 30

    for c_idx, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=HEADER_ROW, column=c_idx, value=str(col).replace("_", " ").title())
        cell.font = _font(bold=True, color=Theme.HEADER_FG, size=10)
        cell.fill = _fill(Theme.HEADER_BG)
        cell.alignment = _align(
            "right" if col_types[col] in ("currency", "float", "integer", "percent") else "center"
        )
        cell.border = _header_bottom_border()

    # ── Data rows ─────────────────────────────────────────────────────────────
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=DATA_START):
        bg = Theme.ROW_ALT if (r_idx - DATA_START) % 2 == 1 else Theme.ROW_DEFAULT
        ws.row_dimensions[r_idx].height = 18

        for c_idx, (value, col) in enumerate(zip(row, df.columns), start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            col_type = col_types[col]
            cell.fill = _fill(bg)
            cell.border = _border()
            cell.font = _font(size=10)
            cell.number_format = _FORMAT_MAP.get(col_type, "@")
            cell.alignment = _align(
                "right" if col_type in ("currency", "float", "integer", "percent") else "left"
            )

    # ── Totals row (numeric columns only) ────────────────────────────────────
    last_data_row = DATA_START + len(df) - 1
    totals_row    = last_data_row + 1
    ws.row_dimensions[totals_row].height = 22
    has_totals = False

    for c_idx, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=totals_row, column=c_idx)
        col_type = col_types[col]
        if col_type in ("currency", "integer", "float"):
            col_letter = get_column_letter(c_idx)
            cell.value = f"=SUM({col_letter}{DATA_START}:{col_letter}{last_data_row})"
            cell.number_format = _FORMAT_MAP[col_type]
            cell.font = _font(bold=True, color=Theme.TOTAL_FG, size=10)
            cell.fill = _fill(Theme.TOTAL_BG)
            cell.alignment = _align("right")
            has_totals = True
        elif c_idx == 1:
            cell.value = "TOTAL"
            cell.font = _font(bold=True, color=Theme.TOTAL_FG, size=10)
            cell.fill = _fill(Theme.TOTAL_BG)
            cell.alignment = _align("left")
        else:
            cell.fill = _fill(Theme.TOTAL_BG)

        thick = Side(style="medium", color=Theme.ACCENT)
        thin  = Side(style="thin",   color=Theme.BORDER_COLOR)
        cell.border = Border(left=thin, right=thin, top=thick, bottom=thick)

    # ── Excel Table (structured reference, auto-filter, sortable) ────────────
    table_ref = (
        f"A{HEADER_ROW}:{last_col_letter}{last_data_row}"
    )
    safe_sheet_name = "".join(c if c.isalnum() else "_" for c in sheet_name)
    table = Table(displayName=f"Table_{safe_sheet_name}", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    # ── Column widths ─────────────────────────────────────────────────────────
    _auto_width(ws, df, offset_col=1)

    # ── Freeze panes (header + title rows) ───────────────────────────────────
    ws.freeze_panes = f"A{DATA_START}"

    # ── Print settings ────────────────────────────────────────────────────────
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.orientation = "landscape"
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)
    ws.print_title_rows = f"$1:${HEADER_ROW}"    # repeat header on every printed page

    # ── Row 1 column A background (covers left of title) ─────────────────────
    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width, 18)

    return ws


# ─── PUBLIC API ───────────────────────────────────────────────────────────────

def csv_to_excel(
    input_path: str,
    output_path: str,
    report_title: str | None = None,
    sheet_name: str = "Data",
    include_summary: bool = False,
    encoding: str = "utf-8-sig",       # handles BOM from Windows Excel exports
    delimiter: str = ",",
) -> Path:
    """
    Convert any CSV file into a professionally formatted Excel workbook.

    Parameters
    ----------
    input_path      : Path to the source CSV file.
    output_path     : Destination .xlsx path.
    report_title    : Custom title shown in the banner (defaults to filename).
    sheet_name      : Name for the main data worksheet.
    include_summary : If True, append an auto-generated summary sheet.
    encoding        : CSV encoding (default handles UTF-8 with/without BOM).
    delimiter       : CSV delimiter (default ',').

    Returns
    -------
    Path to the saved .xlsx file.
    """
    input_path  = Path(input_path)
    output_path = Path(output_path)

    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")
    if not output_path.suffix.lower() == ".xlsx":
        output_path = output_path.with_suffix(".xlsx")

    # ── Load CSV ──────────────────────────────────────────────────────────────
    df = pd.read_csv(
        input_path,
        encoding=encoding,
        sep=delimiter,
        low_memory=False,
    )

    # Light cleaning: strip whitespace from column names and string cells
    df.columns = df.columns.str.strip()
    for col in df.select_dtypes(include=["object", "string"]).columns:
        df[col] = df[col].str.strip()

    # Try to parse columns that look like dates
    for col in df.select_dtypes(include=["object", "string"]).columns:
        if any(k in col.lower() for k in ("date", "time", "created", "updated", "at")):
            try:
                df[col] = pd.to_datetime(df[col], infer_datetime_format=True)
            except Exception:
                pass

    title  = report_title or input_path.stem.replace("_", " ").replace("-", " ").title()
    source = input_path.name

    # ── Build workbook ────────────────────────────────────────────────────────
    wb = Workbook()
    _build_data_sheet(wb, df, sheet_name, title, source)

    if include_summary:
        _build_summary_sheet(wb, df, source)

    wb.save(output_path)
    return output_path


# ─── CLI ENTRY POINT ──────────────────────────────────────────────────────────

def _cli():
    parser = argparse.ArgumentParser(
        description="Convert any CSV to a professional Excel report.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python csv_to_excel_formatter.py --input sales.csv --output sales_report.xlsx
  python csv_to_excel_formatter.py --input data.csv  --output report.xlsx --title "Q1 Dashboard" --summary
  python csv_to_excel_formatter.py --input export.csv --output out.xlsx --sheet "Raw Data" --delimiter ";"
        """
    )
    parser.add_argument("--input",     required=True,  help="Path to source CSV file")
    parser.add_argument("--output",    required=True,  help="Path for output .xlsx file")
    parser.add_argument("--title",     default=None,   help="Report title (defaults to filename)")
    parser.add_argument("--sheet",     default="Data", help="Main sheet name (default: 'Data')")
    parser.add_argument("--summary",   action="store_true", help="Include auto-generated summary sheet")
    parser.add_argument("--encoding",  default="utf-8-sig", help="CSV encoding (default: utf-8-sig)")
    parser.add_argument("--delimiter", default=",",    help="CSV delimiter (default: ',')")
    args = parser.parse_args()

    out = csv_to_excel(
        input_path=args.input,
        output_path=args.output,
        report_title=args.title,
        sheet_name=args.sheet,
        include_summary=args.summary,
        encoding=args.encoding,
        delimiter=args.delimiter,
    )
    print(f"✅  Saved → {out}")


if __name__ == "__main__":
    _cli()
